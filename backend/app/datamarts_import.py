"""Import Data Marts roadmap from DataMarts.xlsx."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import Category, ComponentSubStage, Project, ProjectComponent, Task, TaskStatus, TaskSubStage
from app.services.table_schema import default_table_schema

RU_MONTHS = {
    "янв": 1,
    "февр": 2,
    "мар": 3,
    "апр": 4,
    "май": 5,
    "июн": 6,
    "июл": 7,
    "авг": 8,
    "сент": 9,
    "окт": 10,
    "нояб": 11,
    "дек": 12,
}

PHASE_COLUMNS = [
    "Дата-архитектура (согласованная с Матвеем)",
    "Детальный слой Аналитика",
    "Детальный слой Разработка",
    "Витрина данных Аналитика",
    "Витрина данных Разработка",
    "Витрина данных Публикация",
    "BI Аналитика",
    "BI Разработка",
    "TLA / Оферта",
    "ККД Аналитика",
    "ККД Разработка",
    "Отчетность Разработка",
    "Отчетность PR",
    "Отчетность Внедрение",
]

CATEGORY_COLORS = [
    "#2563eb",
    "#16a34a",
    "#d97706",
    "#7c3aed",
    "#db2777",
    "#0891b2",
    "#ca8a04",
    "#dc2626",
    "#4f46e5",
    "#0d9488",
]

PROJECT_NAME = "Витрины данных"


@dataclass
class ParsedPhase:
    name: str
    due_date: date | None
    is_done: bool
    is_indicative: bool
    note: str


@dataclass
class ParsedRow:
    priority: int | None
    status: TaskStatus
    category: str
    subproduct: str | None
    forms: str | None
    customer: str | None
    data_source: str | None
    platform: str | None
    area: str | None
    contractor: str | None
    desired_quarter: str | None
    attribute_count: str | None
    assignee: str | None
    risks: str | None
    notes: str | None
    extra_info: str | None
    phases: list[ParsedPhase]
    start_date: date | None
    end_date: date | None
    indicative_start: date | None
    indicative_end: date | None
    completion_pct: int
    name: str


def resolve_xlsx_path() -> Path:
    here = Path(__file__).resolve()
    candidates = [
        Path("/app/import-data/DataMarts.xlsx"),
        here.parents[1] / "data" / "DataMarts.xlsx",
    ]
    for depth in (3, 2):
        if depth < len(here.parents):
            candidates.append(here.parents[depth] / "data" / "DataMarts.xlsx")
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError("DataMarts.xlsx not found in import-data or data/")


def _cell_str(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    return text or None


def _parse_ru_date_fragment(fragment: str) -> date | None:
    fragment = fragment.strip()
    if not fragment or fragment.upper() == "DONE":
        return None

    m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", fragment)
    if m:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))

    m = re.search(r"(\d{1,2})\s+([а-яё]+)\.?\s+(\d{4})", fragment, flags=re.I)
    if m:
        day = int(m.group(1))
        mon_key = m.group(2).lower()[:5]
        year = int(m.group(3))
        for prefix, month in RU_MONTHS.items():
            if mon_key.startswith(prefix):
                return date(year, month, day)
    return None


def parse_phase_value(value) -> ParsedPhase | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return ParsedPhase("", value.date(), False, False, value.strftime("%Y-%m-%d"))
    if isinstance(value, date):
        return ParsedPhase("", value, False, False, value.isoformat())

    raw = str(value).strip()
    if not raw or raw == "0":
        return None

    is_indicative = "индикатив" in raw.lower()
    is_done = raw.upper() == "DONE" or re.fullmatch(r"индикатив:DONE", raw, flags=re.I)

    dates: list[date] = []
    for part in re.split(r"[\n;]+", raw):
        cleaned = re.sub(r"^индикатив(\s+с\s+[\d.]+)?:\s*", "", part.strip(), flags=re.I)
        cleaned = re.sub(r"^индикатив\s*", "", cleaned, flags=re.I)
        if cleaned.upper() == "DONE":
            is_done = True
            continue
        parsed = _parse_ru_date_fragment(cleaned)
        if parsed:
            dates.append(parsed)

    due = max(dates) if dates else None
    if is_done and not due:
        return ParsedPhase("", None, True, is_indicative, raw)
    if due or is_done or raw:
        return ParsedPhase("", due, is_done, is_indicative, raw)
    return None


def map_status(value: str | None) -> TaskStatus:
    if not value:
        return TaskStatus.todo
    lower = value.lower()
    if "работе" in lower:
        return TaskStatus.in_progress
    if "пауз" in lower:
        return TaskStatus.blocked
    if "готов" in lower or "done" in lower:
        return TaskStatus.done
    return TaskStatus.todo


def build_usage_name(category: str, subproduct: str | None, data_source: str | None = None) -> str:
    parts = [category]
    if subproduct:
        parts.append(subproduct)
    if data_source:
        parts.append(data_source)
    return " · ".join(parts) if len(parts) > 1 else category


def disambiguate_usage_name(base: str, row: ParsedRow, used_names: set[str]) -> str:
    """Ensure unique task names without numeric suffixes when possible."""
    if base not in used_names:
        return base
    for extra in (row.data_source, row.forms, row.customer):
        if not extra:
            continue
        candidate = f"{base} · {extra}"
        if candidate not in used_names:
            return candidate
    suffix = 2
    while f"{base} ({suffix})" in used_names:
        suffix += 1
    return f"{base} ({suffix})"


def source_key(data_source: str | None, fallback: str) -> str:
    return (data_source or fallback).strip()


def pick_canonical_row(rows: list[ParsedRow]) -> ParsedRow:
    return max(rows, key=lambda r: (len(r.phases), r.completion_pct, bool(r.start_date)))


def create_component_from_row(db: Session, project_id: int, row: ParsedRow) -> ProjectComponent:
    key = source_key(row.data_source, row.name)
    duration = None
    if row.start_date and row.end_date:
        duration = (row.end_date - row.start_date).days + 1
    component = ProjectComponent(
        project_id=project_id,
        name=key,
        data_source=key,
        assignee=row.assignee,
        status=row.status,
        completion_pct=row.completion_pct,
        start_date=row.start_date,
        end_date=row.end_date,
        duration_days=duration,
        indicative_start=row.indicative_start,
        indicative_end=row.indicative_end,
        contractor=row.contractor,
        platform=row.platform,
        notes=row.notes,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(component)
    db.flush()
    for sort_order, phase in enumerate(row.phases):
        db.add(
            ComponentSubStage(
                component_id=component.id,
                name=phase.name,
                sort_order=sort_order,
                is_done=phase.is_done,
                due_date=phase.due_date,
                note=phase.note,
                is_indicative=phase.is_indicative,
            )
        )
    return component


def build_task_name(category: str, data_source: str | None, subproduct: str | None, row_idx: int) -> str:
    parts = [p for p in (data_source, subproduct) if p]
    if parts:
        return " — ".join(parts)
    if category:
        return f"{category} (#{row_idx})"
    return f"Строка {row_idx}"


def compute_task_dates(phases: list[ParsedPhase]) -> tuple[date | None, date | None, date | None, date | None, int]:
    actual_dates: list[date] = []
    indicative_dates: list[date] = []
    done_count = 0
    total = 0

    for phase in phases:
        if not phase.note and not phase.due_date and not phase.is_done:
            continue
        total += 1
        if phase.is_done:
            done_count += 1
        elif phase.due_date:
            if phase.is_indicative:
                indicative_dates.append(phase.due_date)
            else:
                actual_dates.append(phase.due_date)

    start = min(actual_dates) if actual_dates else None
    end = max(actual_dates) if actual_dates else None
    ind_start = min(indicative_dates) if indicative_dates else None
    ind_end = max(indicative_dates) if indicative_dates else None
    pct = round(done_count / total * 100) if total else 0
    return start, end, ind_start, ind_end, pct


def parse_workbook(path: Path) -> list[ParsedRow]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    return parse_worksheet(ws.max_column, ws.max_row, lambda r, c: ws.cell(r, c).value)


def parse_workbook_bytes(content: bytes) -> list[ParsedRow]:
    from io import BytesIO

    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    return parse_worksheet(ws.max_column, ws.max_row, lambda r, c: ws.cell(r, c).value)


def parse_xls_bytes(content: bytes) -> list[ParsedRow]:
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)

    def cell_value(row_idx: int, col_idx: int):
        if row_idx >= sheet.nrows or col_idx >= sheet.ncols:
            return None
        cell = sheet.cell(row_idx, col_idx)
        if cell.ctype == xlrd.XL_CELL_DATE:
            return xlrd.xldate.xldate_as_datetime(cell.value, book.datemode).date()
        return cell.value

    return parse_worksheet(sheet.ncols, sheet.nrows, lambda r, c: cell_value(r - 1, c - 1))


def parse_worksheet(max_column: int, max_row: int, get_cell) -> list[ParsedRow]:
    headers = [_cell_str(get_cell(1, col)) or "" for col in range(1, max_column + 1)]
    header_index = {name: idx + 1 for idx, name in enumerate(headers) if name}

    def col(name: str) -> int | None:
        return header_index.get(name)

    rows: list[ParsedRow] = []
    for row_idx in range(2, max_row + 1):
        category = _cell_str(get_cell(row_idx, col("БВ") or 3)) or ""
        data_source = _cell_str(get_cell(row_idx, col("Источник") or 7))
        subproduct = _cell_str(get_cell(row_idx, col("Субпродукт") or 4))

        if not any([category, data_source, subproduct]):
            continue

        priority_raw = get_cell(row_idx, col("Приоритет") or 1)
        priority = int(priority_raw) if priority_raw not in (None, "") else None

        phases: list[ParsedPhase] = []
        for phase_name in PHASE_COLUMNS:
            phase_col = col(phase_name)
            if not phase_col:
                continue
            parsed = parse_phase_value(get_cell(row_idx, phase_col))
            if parsed:
                phases.append(
                    ParsedPhase(
                        name=phase_name,
                        due_date=parsed.due_date,
                        is_done=parsed.is_done,
                        is_indicative=parsed.is_indicative,
                        note=parsed.note,
                    )
                )

        start, end, ind_start, ind_end, pct = compute_task_dates(phases)
        status = map_status(_cell_str(get_cell(row_idx, col("Статус") or 2)))
        if pct >= 100 and status != TaskStatus.blocked:
            status = TaskStatus.done

        rows.append(
            ParsedRow(
                priority=priority,
                status=status,
                category=category or "Без категории",
                subproduct=subproduct,
                forms=_cell_str(get_cell(row_idx, col("Формы") or 5)),
                customer=_cell_str(get_cell(row_idx, col("Заказчик") or 6)),
                data_source=data_source,
                platform=_cell_str(get_cell(row_idx, col("Площадка") or 8)),
                area=_cell_str(get_cell(row_idx, col("Область") or 9)),
                contractor=_cell_str(get_cell(row_idx, col("Подрядчик") or 12)),
                desired_quarter=_cell_str(get_cell(row_idx, col("Желаемый срок реализации") or 13)),
                attribute_count=_cell_str(get_cell(row_idx, col("Количество атрибутов") or 14)),
                assignee=_cell_str(get_cell(row_idx, col("Команда-исполнитель") or 11)),
                risks=_cell_str(get_cell(row_idx, col("Риски") or 31)),
                notes=_cell_str(get_cell(row_idx, col("Комментарий") or 32)),
                extra_info=_cell_str(get_cell(row_idx, col("Прочая полезная инфомрация") or 36)),
                phases=phases,
                start_date=start,
                end_date=end,
                indicative_start=ind_start,
                indicative_end=ind_end,
                completion_pct=pct,
                name=build_task_name(category, data_source, subproduct, row_idx),
            )
        )

    return rows


def parse_spreadsheet(content: bytes, filename: str) -> list[ParsedRow]:
    ext = Path(filename).suffix.lower()
    if ext == ".xlsx":
        return parse_workbook_bytes(content)
    if ext == ".xls":
        return parse_xls_bytes(content)
    raise ValueError(f"Unsupported spreadsheet format: {ext}")


def import_parsed_rows(
    db: Session,
    parsed_rows: list[ParsedRow],
    project_name: str,
    project_description: str | None = None,
) -> Project:
    if not parsed_rows:
        raise ValueError("File contains no importable rows")

    project = Project(
        name=project_name,
        description=project_description,
        table_schema=default_table_schema("datamarts"),
        created_at=datetime.utcnow(),
    )
    db.add(project)
    db.flush()

    categories_by_name: dict[str, Category] = {}
    for idx, cat_name in enumerate(dict.fromkeys(r.category for r in parsed_rows)):
        cat = Category(
            project_id=project.id,
            name=cat_name,
            color=CATEGORY_COLORS[idx % len(CATEGORY_COLORS)],
            sort_order=idx,
        )
        db.add(cat)
        categories_by_name[cat_name] = cat
    db.flush()

    rows_by_source: dict[str, list[ParsedRow]] = {}
    for row in parsed_rows:
        key = source_key(row.data_source, row.name)
        rows_by_source.setdefault(key, []).append(row)

    components_by_source: dict[str, ProjectComponent] = {}
    for key, group in rows_by_source.items():
        canonical = pick_canonical_row(group)
        components_by_source[key] = create_component_from_row(db, project.id, canonical)

    used_names: set[str] = set()
    for row in parsed_rows:
        key = source_key(row.data_source, row.name)
        component = components_by_source[key]
        name = disambiguate_usage_name(
            build_usage_name(row.category, row.subproduct, row.data_source),
            row,
            used_names,
        )
        used_names.add(name)

        task = Task(
            project_id=project.id,
            category_id=categories_by_name[row.category].id,
            component_id=component.id,
            name=name,
            priority=row.priority,
            subproduct=row.subproduct,
            forms=row.forms,
            customer=row.customer,
            data_source=row.data_source,
            area=row.area,
            desired_quarter=row.desired_quarter,
            attribute_count=row.attribute_count,
            risks=row.risks,
            notes=row.notes,
            extra_info=row.extra_info,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        db.add(task)

    db.commit()
    db.refresh(project)
    return project


def import_datamarts(db: Session, xlsx_path: Path | None = None) -> Project:
    path = xlsx_path or resolve_xlsx_path()
    parsed_rows = parse_workbook(path)

    existing = db.query(Project).filter(Project.name == PROJECT_NAME).first()
    if existing:
        db.delete(existing)
        db.flush()

    return import_parsed_rows(
        db,
        parsed_rows,
        PROJECT_NAME,
        "Реестр витрин данных — импорт из DataMarts.xlsx",
    )
