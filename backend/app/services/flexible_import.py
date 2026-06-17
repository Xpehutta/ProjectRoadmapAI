"""Import spreadsheets and JSON with arbitrary columns."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.datamarts_import import CATEGORY_COLORS, map_status
from app.models import Category, Project, Task, TaskStatus

DATAMARTS_MARKERS = frozenset({"бв", "источник", "субпродукт"})

FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "name": (
        "name",
        "task",
        "task name",
        "title",
        "задача",
        "название",
        "использование",
        "наименование",
        "item",
    ),
    "status": ("status", "статус", "state"),
    "category": ("category", "категория", "бв", "group", "группа", "type", "тип"),
    "priority": ("priority", "приоритет", "pri", "rank"),
    "assignee": (
        "assignee",
        "исполнитель",
        "owner",
        "команда",
        "команда-исполнитель",
        "ответственный",
        "assigned to",
    ),
    "start_date": ("start", "start date", "start_date", "начало", "дата начала"),
    "end_date": ("end", "end date", "end_date", "окончание", "дата окончания", "deadline", "срок"),
    "indicative_start": ("indicative start", "indicative_start", "инд. начало", "индикативное начало"),
    "indicative_end": ("indicative end", "indicative_end", "инд. окончание", "индикативное окончание"),
    "duration_days": ("duration", "duration_days", "длительность", "days"),
    "planned_cost": ("planned cost", "planned_cost", "план. стоимость", "плановая стоимость", "budget"),
    "actual_cost": ("actual cost", "actual_cost", "факт. стоимость", "фактическая стоимость"),
    "planned_effort": ("planned effort", "planned_effort", "план. трудозатраты", "planned hours"),
    "actual_effort": ("actual effort", "actual_effort", "факт. трудозатраты", "actual hours"),
    "notes": ("notes", "comment", "comments", "комментарий", "комментарии", "note"),
    "risks": ("risks", "risk", "риски", "риск"),
    "data_source": ("data source", "data_source", "источник", "source"),
    "subproduct": ("subproduct", "субпродукт"),
    "forms": ("forms", "формы"),
    "customer": ("customer", "заказчик", "client"),
    "platform": ("platform", "площадка"),
    "area": ("area", "область"),
    "contractor": ("contractor", "подрядчик"),
    "completion_pct": ("completion", "completion_pct", "progress", "percent", "%", "выполнение"),
}


@dataclass
class ColumnSpec:
    key: str
    label: str
    col_type: str
    source: str  # "builtin" | "custom"


@dataclass
class FlexibleRow:
    builtin: dict[str, object]
    custom: dict[str, str]
    category: str


def normalize_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_header_key(header: str) -> str:
    return re.sub(r"\s+", " ", header.strip().lower())


def is_datamarts_headers(headers: list[str]) -> bool:
    keys = {normalize_header_key(h) for h in headers if h}
    return DATAMARTS_MARKERS.issubset(keys)


def _cell_str(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _parse_date_value(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    if re.match(r"^\d{4}-\d{2}-\d{2}", text):
        return date.fromisoformat(text[:10])
    m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", text)
    if m:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    return None


def _parse_status_value(value) -> TaskStatus:
    if not value:
        return TaskStatus.todo
    text = str(value).strip().lower()
    if text in TaskStatus.__members__:
        return TaskStatus(text)
    return map_status(str(value))


def map_header_to_field(header: str) -> str | None:
    key = normalize_header_key(header)
    if not key:
        return None
    for field, aliases in FIELD_ALIASES.items():
        if key in aliases:
            return field
    return None


def infer_column_type(field: str | None, sample_values: list) -> str:
    if field == "status":
        return "status"
    if field == "category":
        return "category"
    if field in {"start_date", "end_date", "indicative_start", "indicative_end"}:
        return "date"
    if field in {"priority", "duration_days", "completion_pct"}:
        return "number"
    if field in {"notes", "risks", "extra_info"}:
        return "textarea"
    if any(_parse_date_value(v) for v in sample_values[:5] if v not in (None, "")):
        return "date"
    return "text"


def custom_column_key(index: int) -> str:
    return f"custom_{index}"


def build_column_specs(headers: list[str], rows: list[dict[str, object]]) -> list[ColumnSpec]:
    specs: list[ColumnSpec] = []
    used_fields: set[str] = set()

    for index, header in enumerate(headers):
        label = normalize_header(header)
        if not label:
            continue
        mapped = map_header_to_field(label)
        samples = [row.get(label) for row in rows[:20]]
        if mapped and mapped not in used_fields:
            used_fields.add(mapped)
            api_key = "category_id" if mapped == "category" else mapped
            specs.append(
                ColumnSpec(
                    key=api_key,
                    label=label,
                    col_type=infer_column_type(mapped, samples),
                    source="builtin",
                )
            )
        else:
            specs.append(
                ColumnSpec(
                    key=custom_column_key(index),
                    label=label,
                    col_type=infer_column_type(None, samples),
                    source="custom",
                )
            )

    spec_keys = {s.key for s in specs}
    if "status" not in spec_keys:
        specs.insert(0, ColumnSpec(key="status", label="Статус", col_type="status", source="builtin"))
    if "name" not in spec_keys:
        insert_at = 1 if specs and specs[0].key == "status" else 0
        specs.insert(insert_at, ColumnSpec(key="name", label="Задача", col_type="text", source="builtin"))
    return specs


def schema_to_json(specs: list[ColumnSpec]) -> list[dict]:
    return [
        {"key": s.key, "label": s.label, "type": s.col_type, "source": s.source}
        for s in specs
    ]


def _read_xlsx_sheet(content: bytes) -> tuple[list[str], list[list]]:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    headers = [_cell_str(ws.cell(1, col).value) or "" for col in range(1, ws.max_column + 1)]
    while headers and not headers[-1]:
        headers.pop()
    rows: list[list] = []
    for row_idx in range(2, ws.max_row + 1):
        rows.append([ws.cell(row_idx, col).value for col in range(1, len(headers) + 1)])
    return headers, rows


def _read_xls_sheet(content: bytes) -> tuple[list[str], list[list]]:
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    headers = [_cell_str(sheet.cell_value(0, col)) or "" for col in range(sheet.ncols)]
    while headers and not headers[-1]:
        headers.pop()
    rows: list[list] = []
    for row_idx in range(1, sheet.nrows):
        row: list = []
        for col_idx in range(len(headers)):
            if col_idx >= sheet.ncols:
                row.append(None)
                continue
            cell = sheet.cell(row_idx, col_idx)
            if cell.ctype == xlrd.XL_CELL_DATE:
                row.append(xlrd.xldate.xldate_as_datetime(cell.value, book.datemode).date())
            else:
                row.append(cell.value)
        rows.append(row)
    return headers, rows


def read_spreadsheet_sheet(content: bytes, filename: str) -> tuple[list[str], list[list]]:
    ext = Path(filename).suffix.lower()
    if ext == ".xlsx":
        return _read_xlsx_sheet(content)
    if ext == ".xls":
        return _read_xls_sheet(content)
    raise ValueError(f"Unsupported spreadsheet format: {ext}")


def peek_spreadsheet_headers(content: bytes, filename: str) -> list[str]:
    headers, _ = read_spreadsheet_sheet(content, filename)
    return [h for h in headers if normalize_header(h)]


def rows_from_sheet(headers: list[str], raw_rows: list[list]) -> list[dict[str, object]]:
    parsed: list[dict[str, object]] = []
    for raw in raw_rows:
        if not any(v not in (None, "") for v in raw):
            continue
        row: dict[str, object] = {}
        for index, header in enumerate(headers):
            label = normalize_header(header)
            if not label:
                continue
            row[label] = raw[index] if index < len(raw) else None
        if row:
            parsed.append(row)
    return parsed


def parse_flexible_row(
    row: dict[str, object],
    specs: list[ColumnSpec],
    row_idx: int,
) -> FlexibleRow | None:
    builtin: dict[str, object] = {}
    custom: dict[str, str] = {}
    category_name = "Без категории"

    for spec in specs:
        value = row.get(spec.label)
        if value in (None, ""):
            continue
        if spec.source == "custom":
            text = _cell_str(value)
            if text:
                custom[spec.key] = text
            continue

        field = "category" if spec.key == "category_id" else spec.key
        if field == "category":
            category_name = str(value).strip() or category_name
            continue
        if field == "status":
            builtin["status"] = _parse_status_value(value)
            continue
        if field in {"start_date", "end_date", "indicative_start", "indicative_end"}:
            parsed_date = _parse_date_value(value)
            if parsed_date:
                builtin[field] = parsed_date
            continue
        if field in {"priority", "duration_days", "completion_pct"}:
            try:
                builtin[field] = int(float(str(value).replace(",", ".")))
            except (TypeError, ValueError):
                pass
            continue
        text = _cell_str(value)
        if text:
            builtin[field] = text

    if not builtin.get("name"):
        for spec in specs:
            if spec.source == "custom":
                text = custom.get(spec.key)
                if text:
                    builtin["name"] = text
                    break
        if not builtin.get("name"):
            parts = [str(v) for v in row.values() if v not in (None, "")]
            if parts:
                builtin["name"] = parts[0] if len(parts) == 1 else " · ".join(parts[:2])
            else:
                return None

    return FlexibleRow(builtin=builtin, custom=custom, category=category_name)


def _create_task_from_flexible_row(
    project_id: int,
    category_id: int,
    parsed: FlexibleRow,
    row_idx: int,
) -> Task:
    start = parsed.builtin.get("start_date")
    end = parsed.builtin.get("end_date")
    duration = parsed.builtin.get("duration_days")
    if start and end and not duration:
        duration = (end - start).days + 1  # type: ignore[operator]

    task = Task(
        project_id=project_id,
        category_id=category_id,
        name=str(parsed.builtin.get("name") or f"Строка {row_idx}"),
        status=parsed.builtin.get("status") or TaskStatus.todo,
        completion_pct=int(parsed.builtin.get("completion_pct") or 0),
        priority=parsed.builtin.get("priority"),  # type: ignore[arg-type]
        assignee=parsed.builtin.get("assignee"),  # type: ignore[arg-type]
        start_date=start,  # type: ignore[arg-type]
        end_date=end,  # type: ignore[arg-type]
        duration_days=duration,  # type: ignore[arg-type]
        indicative_start=parsed.builtin.get("indicative_start"),  # type: ignore[arg-type]
        indicative_end=parsed.builtin.get("indicative_end"),  # type: ignore[arg-type]
        planned_cost=parsed.builtin.get("planned_cost"),  # type: ignore[arg-type]
        actual_cost=parsed.builtin.get("actual_cost"),  # type: ignore[arg-type]
        planned_effort=parsed.builtin.get("planned_effort"),  # type: ignore[arg-type]
        actual_effort=parsed.builtin.get("actual_effort"),  # type: ignore[arg-type]
        subproduct=parsed.builtin.get("subproduct"),  # type: ignore[arg-type]
        forms=parsed.builtin.get("forms"),  # type: ignore[arg-type]
        customer=parsed.builtin.get("customer"),  # type: ignore[arg-type]
        data_source=parsed.builtin.get("data_source"),  # type: ignore[arg-type]
        platform=parsed.builtin.get("platform"),  # type: ignore[arg-type]
        area=parsed.builtin.get("area"),  # type: ignore[arg-type]
        contractor=parsed.builtin.get("contractor"),  # type: ignore[arg-type]
        notes=parsed.builtin.get("notes"),  # type: ignore[arg-type]
        risks=parsed.builtin.get("risks"),  # type: ignore[arg-type]
        custom_fields=parsed.custom or None,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    return task


def import_flexible_sheet(
    db: Session,
    *,
    headers: list[str],
    raw_rows: list[list],
    project_name: str,
    project_description: str | None = None,
) -> Project:
    dict_rows = rows_from_sheet(headers, raw_rows)
    if not dict_rows:
        raise ValueError("File contains no data rows")

    specs = build_column_specs(headers, dict_rows)
    project = Project(
        name=project_name,
        description=project_description,
        table_schema=schema_to_json(specs),
        created_at=datetime.utcnow(),
    )
    db.add(project)
    db.flush()

    categories: dict[str, Category] = {}

    def ensure_category(name: str) -> Category:
        if name in categories:
            return categories[name]
        idx = len(categories)
        cat = Category(
            project_id=project.id,
            name=name,
            color=CATEGORY_COLORS[idx % len(CATEGORY_COLORS)],
            sort_order=idx,
        )
        db.add(cat)
        db.flush()
        categories[name] = cat
        return cat

    for row_idx, row in enumerate(dict_rows, start=1):
        parsed = parse_flexible_row(row, specs, row_idx)
        if not parsed:
            continue
        category = ensure_category(parsed.category)
        db.add(_create_task_from_flexible_row(project.id, category.id, parsed, row_idx))

    db.commit()
    db.refresh(project)
    return project


def import_flexible_spreadsheet(
    db: Session,
    content: bytes,
    filename: str,
    project_name: str,
    project_description: str | None = None,
) -> Project:
    headers, raw_rows = read_spreadsheet_sheet(content, filename)
    return import_flexible_sheet(
        db,
        headers=headers,
        raw_rows=raw_rows,
        project_name=project_name,
        project_description=project_description,
    )


def import_flexible_json(
    db: Session,
    payload,
    project_name: str,
    project_description: str | None = None,
) -> Project:
    columns = None
    rows = None
    name = project_name
    description = project_description

    if isinstance(payload, list):
        rows = payload
    elif isinstance(payload, dict):
        if payload.get("format") in ("flexible", "roadmap"):
            columns = payload.get("columns")
            rows = payload.get("rows") or payload.get("tasks") or payload.get("data")
            name = str(payload.get("name") or project_name)
            description = payload.get("description") or project_description
        else:
            rows = payload.get("rows") or payload.get("tasks") or payload.get("data")
            columns = payload.get("columns")
            name = str(payload.get("name") or project_name)
            description = payload.get("description") or project_description
    else:
        raise ValueError("JSON must be an object or array")

    if not isinstance(rows, list) or not rows:
        raise ValueError("JSON must contain at least one row")

    dict_rows: list[dict[str, object]] = []
    for item in rows:
        if isinstance(item, dict):
            dict_rows.append({normalize_header(k): v for k, v in item.items() if normalize_header(k)})

    if not dict_rows:
        raise ValueError("No valid rows found in JSON")

    if columns:
        specs = [
            ColumnSpec(
                key=str(c.get("key") or custom_column_key(i)),
                label=str(c.get("label") or c.get("key") or f"Column {i}"),
                col_type=str(c.get("type") or "text"),
                source=str(c.get("source") or "custom"),
            )
            for i, c in enumerate(columns)
            if isinstance(c, dict)
        ]
    else:
        headers = list(dict.fromkeys(key for row in dict_rows for key in row.keys()))
        specs = build_column_specs(headers, dict_rows)

    project = Project(
        name=name,
        description=description,
        table_schema=schema_to_json(specs),
        created_at=datetime.utcnow(),
    )
    db.add(project)
    db.flush()

    categories: dict[str, Category] = {}

    def ensure_category(cat_name: str) -> Category:
        if cat_name in categories:
            return categories[cat_name]
        idx = len(categories)
        cat = Category(
            project_id=project.id,
            name=cat_name,
            color=CATEGORY_COLORS[idx % len(CATEGORY_COLORS)],
            sort_order=idx,
        )
        db.add(cat)
        db.flush()
        categories[cat_name] = cat
        return cat

    for row_idx, row in enumerate(dict_rows, start=1):
        parsed = parse_flexible_row(row, specs, row_idx)
        if not parsed:
            continue
        category = ensure_category(parsed.category)
        db.add(_create_task_from_flexible_row(project.id, category.id, parsed, row_idx))

    db.commit()
    db.refresh(project)
    return project
